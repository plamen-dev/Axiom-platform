"""Discipline-based extraction engine for inventory runs.

Splits inventory elements by discipline, writes per-discipline artifacts
(parquet, CSV, XLSX, markdown), checkpoints progress, and generates
a root-level discipline run summary.
"""

from __future__ import annotations

import csv
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from axiom_core.inventory.discipline import (
    DISCIPLINES,
    classify_elements,
)
from axiom_core.inventory.storage import (
    write_elements_parquet,
    write_parameters_parquet,
)


def _write_checkpoint(
    path: Path,
    run_id: str,
    discipline: str,
    status: str,
    started_at: str,
    completed_at: str = "",
    elapsed_seconds: float = 0,
    element_count: int = 0,
    type_count: int = 0,
    parameter_count: int = 0,
    category_count: int = 0,
    output_folder: str = "",
    error: str = "",
    warning: str = "",
) -> None:
    """Append a checkpoint entry to the JSONL file."""
    entry = {
        "run_id": run_id,
        "discipline": discipline,
        "status": status,
        "started_at": started_at,
        "completed_at": completed_at,
        "elapsed_seconds": elapsed_seconds,
        "element_count": element_count,
        "type_count": type_count,
        "parameter_count": parameter_count,
        "category_count": category_count,
        "output_folder": output_folder,
        "error": error,
        "warning": warning,
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry, default=str) + "\n")


def _write_discipline_csv(
    elements: list[dict], path: Path, run_id: str, source_model: str,
) -> Path:
    """Write a flat CSV of elements for a discipline chunk."""
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "element_id", "unique_id", "category", "class_name", "name",
        "family_name", "type_name", "level_name", "is_type",
        "parameter_count", "discipline", "discipline_reason",
        "classification_confidence", "run_id", "source_model",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for elem in elements:
            row = {
                "element_id": elem.get("ElementId", 0),
                "unique_id": elem.get("UniqueId", ""),
                "category": elem.get("Category", ""),
                "class_name": elem.get("ClassName", ""),
                "name": elem.get("Name", ""),
                "family_name": elem.get("FamilyName", ""),
                "type_name": elem.get("TypeName", ""),
                "level_name": elem.get("LevelName", ""),
                "is_type": elem.get("IsType", False),
                "parameter_count": len(elem.get("Parameters", [])),
                "discipline": elem.get("discipline", ""),
                "discipline_reason": elem.get("discipline_reason", ""),
                "classification_confidence": elem.get("classification_confidence", ""),
                "run_id": run_id,
                "source_model": source_model,
            }
            writer.writerow(row)
    return path


def _write_discipline_xlsx(
    elements: list[dict],
    path: Path,
    discipline: str,
    run_id: str,
) -> Path:
    """Write a formatted Excel summary for a discipline chunk."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # ── Elements sheet ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Elements"

    headers = [
        "element_id", "category", "name", "family_name", "type_name",
        "is_type", "parameter_count", "discipline_reason",
        "classification_confidence",
    ]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, elem in enumerate(elements, 2):
        vals = [
            elem.get("ElementId", 0),
            elem.get("Category", ""),
            elem.get("Name", ""),
            elem.get("FamilyName", ""),
            elem.get("TypeName", ""),
            str(elem.get("IsType", False)),
            len(elem.get("Parameters", [])),
            elem.get("discipline_reason", ""),
            elem.get("classification_confidence", ""),
        ]
        for col_idx, val in enumerate(vals, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    ws.freeze_panes = "A2"
    for col_idx, h in enumerate(headers, 1):
        max_len = len(h)
        for row_idx in range(2, min(len(elements) + 2, 102)):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, min(len(cell_val), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    # ── Summary sheet ────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")

    total = len(elements)
    instances = sum(1 for e in elements if not e.get("IsType", False))
    types = sum(1 for e in elements if e.get("IsType", False))
    params = sum(len(e.get("Parameters", [])) for e in elements)
    cat_counts = Counter(e.get("Category", "(No Category)") for e in elements)

    summary_rows = [
        ("Metric", "Value"),
        ("Discipline", discipline),
        ("Run ID", run_id),
        ("Total elements", total),
        ("Instances", instances),
        ("Types", types),
        ("Parameters", params),
        ("Categories", len(cat_counts)),
    ]
    for row_idx, (label, value) in enumerate(summary_rows, 1):
        cell_l = ws_sum.cell(row=row_idx, column=1, value=label)
        cell_v = ws_sum.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            cell_l.font = header_font
            cell_l.fill = header_fill
            cell_v.font = header_font
            cell_v.fill = header_fill

    # Top 20 categories
    start = len(summary_rows) + 2
    ws_sum.cell(row=start, column=1, value="Top Categories").font = Font(bold=True)
    ws_sum.cell(row=start + 1, column=1, value="Category").font = header_font
    ws_sum.cell(row=start + 1, column=2, value="Count").font = header_font
    for i, (cat, cnt) in enumerate(cat_counts.most_common(20)):
        ws_sum.cell(row=start + 2 + i, column=1, value=cat)
        ws_sum.cell(row=start + 2 + i, column=2, value=cnt)

    ws_sum.column_dimensions["A"].width = 35
    ws_sum.column_dimensions["B"].width = 20
    ws_sum.freeze_panes = "A2"

    wb.save(str(path))
    return path


def _write_discipline_summary_md(
    elements: list[dict],
    path: Path,
    discipline: str,
    run_id: str,
) -> Path:
    """Write a markdown summary for a single discipline chunk."""
    path.parent.mkdir(parents=True, exist_ok=True)

    total = len(elements)
    instances = sum(1 for e in elements if not e.get("IsType", False))
    types = sum(1 for e in elements if e.get("IsType", False))
    params = sum(len(e.get("Parameters", [])) for e in elements)
    cat_counts = Counter(e.get("Category", "(No Category)") for e in elements)
    confidence_counts = Counter(e.get("classification_confidence", "unknown") for e in elements)

    lines = [
        f"# {discipline} Inventory Summary",
        "",
        f"**Run ID:** {run_id}",
        "",
        "## Counts",
        "",
        "| Metric | Count |",
        "|--------|-------|",
        f"| Total elements | {total} |",
        f"| Instances | {instances} |",
        f"| Types | {types} |",
        f"| Parameters | {params} |",
        f"| Categories | {len(cat_counts)} |",
        "",
    ]

    if cat_counts:
        lines.extend(["## Top Categories", ""])
        lines.extend(["| Category | Count |", "|----------|-------|"])
        for cat, cnt in cat_counts.most_common(20):
            lines.append(f"| {cat} | {cnt} |")
        lines.append("")

    # Top families/types
    fam_counts = Counter(
        f"{e.get('FamilyName', '')} : {e.get('TypeName', '')}"
        for e in elements if e.get("FamilyName")
    )
    if fam_counts:
        lines.extend(["## Top Families/Types", ""])
        lines.extend(["| Family : Type | Count |", "|---------------|-------|"])
        for fam, cnt in fam_counts.most_common(20):
            lines.append(f"| {fam} | {cnt} |")
        lines.append("")

    # Confidence breakdown
    if confidence_counts:
        lines.extend(["## Classification Confidence", ""])
        lines.extend(["| Confidence | Count |", "|------------|-------|"])
        for conf, cnt in confidence_counts.most_common():
            lines.append(f"| {conf} | {cnt} |")
        lines.append("")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return path


def _write_discipline_metadata(
    elements: list[dict],
    path: Path,
    discipline: str,
    run_id: str,
    source_model: str,
    elapsed_seconds: float,
) -> Path:
    """Write run_metadata.json for a discipline chunk."""
    path.parent.mkdir(parents=True, exist_ok=True)
    instances = sum(1 for e in elements if not e.get("IsType", False))
    types = sum(1 for e in elements if e.get("IsType", False))
    params = sum(len(e.get("Parameters", [])) for e in elements)
    cat_counts = dict(Counter(e.get("Category", "(No Category)") for e in elements).most_common())

    meta = {
        "run_id": run_id,
        "discipline": discipline,
        "source_model": source_model,
        "instance_count": instances,
        "type_count": types,
        "parameter_count": params,
        "element_count": len(elements),
        "category_counts": cat_counts,
        "elapsed_seconds": elapsed_seconds,
    }
    path.write_text(json.dumps(meta, indent=2), encoding="utf-8")
    return path


def export_discipline_chunk(
    elements: list[dict],
    discipline_dir: Path,
    discipline: str,
    run_id: str,
    source_model: str,
    elapsed_seconds: float = 0,
) -> dict[str, Path]:
    """Write all artifacts for a single discipline chunk."""
    discipline_dir.mkdir(parents=True, exist_ok=True)
    paths: dict[str, Path] = {}

    paths["elements_parquet"] = write_elements_parquet(
        elements, discipline_dir / "elements.parquet",
        run_id=run_id, source_model=source_model,
    )
    paths["parameters_parquet"] = write_parameters_parquet(
        elements, discipline_dir / "parameters.parquet",
        run_id=run_id,
    )
    paths["elements_csv"] = _write_discipline_csv(
        elements, discipline_dir / "elements.csv",
        run_id=run_id, source_model=source_model,
    )
    paths["xlsx"] = _write_discipline_xlsx(
        elements, discipline_dir / "inventory_summary.xlsx",
        discipline=discipline, run_id=run_id,
    )
    paths["summary_md"] = _write_discipline_summary_md(
        elements, discipline_dir / "inventory_summary.md",
        discipline=discipline, run_id=run_id,
    )
    paths["metadata"] = _write_discipline_metadata(
        elements, discipline_dir / "run_metadata.json",
        discipline=discipline, run_id=run_id,
        source_model=source_model, elapsed_seconds=elapsed_seconds,
    )

    return paths


def _write_root_summary_md(
    discipline_stats: dict[str, dict],
    run_dir: Path,
    run_id: str,
    source_model: str,
    total_elapsed: float,
    empty_input: bool = False,
) -> Path:
    """Write the root discipline_run_summary.md."""
    path = run_dir / "discipline_run_summary.md"

    total_elements = sum(s.get("element_count", 0) for s in discipline_stats.values())
    total_types = sum(s.get("type_count", 0) for s in discipline_stats.values())
    total_params = sum(s.get("parameter_count", 0) for s in discipline_stats.values())
    total_cats = sum(s.get("category_count", 0) for s in discipline_stats.values())
    unknown_count = sum(
        s.get("element_count", 0) for d, s in discipline_stats.items()
        if d == "Other" and s.get("status") == "SUCCESS"
    )

    failed = [d for d, s in discipline_stats.items() if s.get("status") == "FAILED"]
    largest_disc = max(
        discipline_stats.items(),
        key=lambda x: x[1].get("element_count", 0),
        default=("", {}),
    )

    lines = [
        f"# Discipline Run Summary: {run_id}",
        "",
        f"**Source model:** {source_model}",
        f"**Total elapsed:** {total_elapsed:.1f}s",
        "",
    ]

    if empty_input:
        lines.extend([
            "## WARNING",
            "",
            ("Input inventory JSON contains no element-level records. "
             "Discipline split requires full-detail inventory export. "
             "Summary-mode exports cannot be classified by discipline."),
            "",
        ])

    lines.extend([
        "## Totals",
        "",
        "| Metric | Count |",
        "|--------|-------|",
        f"| Total elements | {total_elements} |",
        f"| Total types | {total_types} |",
        f"| Total parameters | {total_params} |",
        f"| Total categories | {total_cats} |",
        f"| Unknown/Other | {unknown_count} |",
        f"| Largest discipline | {largest_disc[0]} ({largest_disc[1].get('element_count', 0)} elements) |",
        "",
        "## By Discipline",
        "",
        "| Discipline | Elements | Types | Parameters | Categories | Elapsed | Status |",
        "|------------|----------|-------|------------|------------|---------|--------|",
    ])

    for disc in DISCIPLINES:
        s = discipline_stats.get(disc, {})
        lines.append(
            f"| {disc} | {s.get('element_count', 0)} | {s.get('type_count', 0)} "
            f"| {s.get('parameter_count', 0)} | {s.get('category_count', 0)} "
            f"| {s.get('elapsed_seconds', 0):.1f}s | {s.get('status', 'SKIPPED')} |"
        )
    lines.append("")

    if failed:
        lines.extend(["## Failed Disciplines", ""])
        for d in failed:
            err = discipline_stats[d].get("error", "unknown error")
            lines.append(f"- **{d}**: {err}")
        lines.append("")

    # Output paths
    if not empty_input:
        lines.extend(["## Output Folders", ""])
        for disc in DISCIPLINES:
            s = discipline_stats.get(disc, {})
            if s.get("output_folder"):
                lines.append(f"- **{disc}:** `{s['output_folder']}`")
        lines.append("")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return path


def _write_root_summary_xlsx(
    discipline_stats: dict[str, dict],
    run_dir: Path,
    run_id: str,
) -> Path:
    """Write root discipline_run_summary.xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    path = run_dir / "discipline_run_summary.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Discipline Summary"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    headers = [
        "Discipline", "Elements", "Types", "Parameters",
        "Categories", "Elapsed (s)", "Status",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row, disc in enumerate(DISCIPLINES, 2):
        s = discipline_stats.get(disc, {})
        ws.cell(row=row, column=1, value=disc)
        ws.cell(row=row, column=2, value=s.get("element_count", 0))
        ws.cell(row=row, column=3, value=s.get("type_count", 0))
        ws.cell(row=row, column=4, value=s.get("parameter_count", 0))
        ws.cell(row=row, column=5, value=s.get("category_count", 0))
        ws.cell(row=row, column=6, value=round(s.get("elapsed_seconds", 0), 1))
        ws.cell(row=row, column=7, value=s.get("status", "SKIPPED"))

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 20
    for col in range(2, 8):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = 15

    wb.save(str(path))
    return path


def _write_root_metadata(
    discipline_stats: dict[str, dict],
    run_dir: Path,
    run_id: str,
    source_model: str,
    total_elapsed: float,
    empty_input: bool = False,
) -> Path:
    """Write root run_metadata.json."""
    path = run_dir / "run_metadata.json"
    meta: dict = {
        "run_id": run_id,
        "source_model": source_model,
        "chunk_by": "discipline",
        "total_elapsed_seconds": total_elapsed,
        "disciplines": {
            d: {
                "element_count": s.get("element_count", 0),
                "type_count": s.get("type_count", 0),
                "parameter_count": s.get("parameter_count", 0),
                "status": s.get("status", "SKIPPED"),
            }
            for d, s in discipline_stats.items()
        },
    }
    if empty_input:
        meta["warning"] = (
            "Input inventory JSON contains no element-level records. "
            "Discipline split requires full-detail inventory export. "
            "Summary-mode exports cannot be classified by discipline."
        )
    path.write_text(json.dumps(meta, indent=2), encoding="utf-8")
    return path


def run_discipline_extraction(
    elements: list[dict],
    output_dir: Path,
    run_id: str,
    source_model: str = "",
    discipline_filter: str | None = None,
) -> dict[str, Path]:
    """Run discipline-based extraction on a set of inventory elements.

    Classifies all elements, then writes per-discipline artifacts and
    a root summary.  If discipline_filter is set, only that discipline
    is exported.

    Returns dict of key output paths.  If the input elements list is
    empty (e.g. from a summary-mode scan), a warning is recorded in
    the summary, metadata, and returned paths dict.
    """
    import time

    run_dir = output_dir / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    checkpoint_path = run_dir / "discipline_checkpoints.jsonl"

    empty_input = len(elements) == 0

    # Classify
    buckets = classify_elements(elements)

    # Filter to single discipline if requested
    if discipline_filter:
        if discipline_filter not in buckets:
            buckets = {discipline_filter: []}
        else:
            buckets = {discipline_filter: buckets[discipline_filter]}

    discipline_stats: dict[str, dict] = {}
    paths: dict[str, Path] = {}
    total_start = time.monotonic()

    for discipline in DISCIPLINES:
        if discipline_filter and discipline != discipline_filter:
            continue

        disc_elements = buckets.get(discipline, [])
        started_at = datetime.now(timezone.utc).isoformat()

        _write_checkpoint(
            checkpoint_path, run_id, discipline, "STARTED",
            started_at=started_at,
        )

        disc_start = time.monotonic()

        try:
            disc_dir = run_dir / discipline
            disc_dir.mkdir(parents=True, exist_ok=True)

            types = sum(1 for e in disc_elements if e.get("IsType", False))
            params = sum(len(e.get("Parameters", [])) for e in disc_elements)
            cats = len(set(e.get("Category", "") for e in disc_elements))

            elapsed = time.monotonic() - disc_start

            if disc_elements:
                export_discipline_chunk(
                    disc_elements, disc_dir, discipline,
                    run_id=run_id, source_model=source_model,
                    elapsed_seconds=elapsed,
                )

            elapsed = time.monotonic() - disc_start
            completed_at = datetime.now(timezone.utc).isoformat()

            stats = {
                "status": "SUCCESS",
                "element_count": len(disc_elements),
                "type_count": types,
                "parameter_count": params,
                "category_count": cats,
                "elapsed_seconds": elapsed,
                "output_folder": str(disc_dir),
            }
            discipline_stats[discipline] = stats

            _write_checkpoint(
                checkpoint_path, run_id, discipline, "SUCCESS",
                started_at=started_at, completed_at=completed_at,
                elapsed_seconds=elapsed,
                element_count=len(disc_elements), type_count=types,
                parameter_count=params, category_count=cats,
                output_folder=str(disc_dir),
            )

        except Exception as exc:
            elapsed = time.monotonic() - disc_start
            completed_at = datetime.now(timezone.utc).isoformat()

            stats = {
                "status": "FAILED",
                "element_count": 0,
                "type_count": 0,
                "parameter_count": 0,
                "category_count": 0,
                "elapsed_seconds": elapsed,
                "error": str(exc),
                "output_folder": "",
            }
            discipline_stats[discipline] = stats

            _write_checkpoint(
                checkpoint_path, run_id, discipline, "FAILED",
                started_at=started_at, completed_at=completed_at,
                elapsed_seconds=elapsed, error=str(exc),
            )

    total_elapsed = time.monotonic() - total_start

    # Root summary outputs
    paths["checkpoint"] = checkpoint_path
    paths["root_summary_md"] = _write_root_summary_md(
        discipline_stats, run_dir, run_id, source_model, total_elapsed,
        empty_input=empty_input,
    )
    paths["root_summary_xlsx"] = _write_root_summary_xlsx(
        discipline_stats, run_dir, run_id,
    )
    paths["root_metadata"] = _write_root_metadata(
        discipline_stats, run_dir, run_id, source_model, total_elapsed,
        empty_input=empty_input,
    )
    if empty_input:
        paths["warning"] = "empty_input"

    return paths
