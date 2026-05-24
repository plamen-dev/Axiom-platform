"""Adaptive extraction planner for inventory runs.

Uses summary-mode category counts to produce an extraction plan that
groups small categories by discipline, isolates large categories, and
chunks very large categories into manageable batches — all before any
full-detail extraction runs.

Default thresholds (configurable via CLI flags):
  - max_group_elements: 5000  — max elements in a discipline group job
  - isolate_category_threshold: 3000  — isolate a category if it exceeds this
  - max_category_chunk_elements: 5000  — max elements per chunk of a large category
"""

from __future__ import annotations

import json
import math
from dataclasses import asdict, dataclass, field
from pathlib import Path

from axiom_core.inventory.discipline import (
    DISCIPLINES,
    _get_bic_lookup,
    _get_keyword_lookup,
)

# ── Default thresholds ───────────────────────────────────────────────

DEFAULT_MAX_GROUP_ELEMENTS = 5000
DEFAULT_ISOLATE_CATEGORY_THRESHOLD = 3000
DEFAULT_MAX_CATEGORY_CHUNK_ELEMENTS = 5000


@dataclass
class ExtractionJob:
    """A single planned extraction job."""

    plan_id: str = ""
    sequence_number: int = 0
    discipline: str = ""
    extraction_scope: str = ""
    categories: list[str] = field(default_factory=list)
    estimated_element_count: int = 0
    strategy: str = ""
    reason: str = ""
    expected_prompt: str = ""
    risk_level: str = "low"
    chunk_index: int | None = None
    total_chunks: int | None = None


@dataclass
class ExtractionPlan:
    """Full adaptive extraction plan for a model."""

    run_id: str = ""
    source_model: str = ""
    total_instance_count: int = 0
    total_type_count: int = 0
    total_category_count: int = 0
    thresholds: dict = field(default_factory=dict)
    jobs: list[ExtractionJob] = field(default_factory=list)
    discipline_totals: dict[str, int] = field(default_factory=dict)
    categories_by_count: list[tuple[str, int]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _classify_category_to_discipline(category_name: str) -> str:
    """Map a category name to a discipline using the mapping JSON."""
    bic_lookup = _get_bic_lookup()
    keyword_lookup = _get_keyword_lookup()

    # Try BIC lookup (category names in summary-mode may match BIC keys
    # or may be plain names — try both)
    disc = bic_lookup.get(category_name)
    if disc:
        return disc

    # Keyword fallback on category name
    cat_lower = category_name.lower()
    for keyword, disc in keyword_lookup:
        if keyword.lower() in cat_lower:
            return disc

    return "Other"


def build_extraction_plan(
    category_counts: dict[str, int],
    run_id: str = "",
    source_model: str = "",
    total_instance_count: int = 0,
    total_type_count: int = 0,
    max_group_elements: int = DEFAULT_MAX_GROUP_ELEMENTS,
    isolate_category_threshold: int = DEFAULT_ISOLATE_CATEGORY_THRESHOLD,
    max_category_chunk_elements: int = DEFAULT_MAX_CATEGORY_CHUNK_ELEMENTS,
) -> ExtractionPlan:
    """Build an adaptive extraction plan from summary-mode category counts.

    Steps:
      1. Classify each category into a discipline.
      2. Within each discipline, separate large categories from small ones.
      3. Large categories (>= isolate_category_threshold) are isolated.
      4. Very large isolated categories (> max_category_chunk_elements) are chunked.
      5. Small categories are grouped by discipline up to max_group_elements.
      6. Groups that exceed max_group_elements are split into sub-groups.
    """
    plan = ExtractionPlan(
        run_id=run_id,
        source_model=source_model,
        total_instance_count=total_instance_count,
        total_type_count=total_type_count,
        total_category_count=len(category_counts),
        thresholds={
            "max_group_elements": max_group_elements,
            "isolate_category_threshold": isolate_category_threshold,
            "max_category_chunk_elements": max_category_chunk_elements,
        },
    )

    # Sort categories by count descending
    sorted_cats = sorted(category_counts.items(), key=lambda x: -x[1])
    plan.categories_by_count = sorted_cats

    # Classify categories into disciplines
    disc_categories: dict[str, list[tuple[str, int]]] = {d: [] for d in DISCIPLINES}
    for cat, count in sorted_cats:
        disc = _classify_category_to_discipline(cat)
        disc_categories[disc].append((cat, count))

    plan.discipline_totals = {
        d: sum(c for _, c in cats) for d, cats in disc_categories.items()
    }

    seq = 1
    jobs: list[ExtractionJob] = []

    for discipline in DISCIPLINES:
        cats = disc_categories[discipline]
        if not cats:
            continue

        large_cats: list[tuple[str, int]] = []
        small_cats: list[tuple[str, int]] = []

        for cat, count in cats:
            if count >= isolate_category_threshold:
                large_cats.append((cat, count))
            else:
                small_cats.append((cat, count))

        # Handle large categories — isolate or chunk
        for cat, count in large_cats:
            if count > max_category_chunk_elements:
                n_chunks = math.ceil(count / max_category_chunk_elements)
                chunk_size = math.ceil(count / n_chunks)
                for ci in range(n_chunks):
                    est = min(chunk_size, count - ci * chunk_size)
                    job = ExtractionJob(
                        plan_id=f"{run_id}_job_{seq:03d}",
                        sequence_number=seq,
                        discipline=discipline,
                        extraction_scope=f"{cat} (chunk {ci + 1}/{n_chunks})",
                        categories=[cat],
                        estimated_element_count=est,
                        strategy="category_chunk",
                        reason=(
                            f"{cat} has {count} elements, exceeds "
                            f"max_category_chunk_elements ({max_category_chunk_elements})"
                        ),
                        expected_prompt=f"Run InventoryModel for {cat}",
                        risk_level="medium" if count < 10000 else "high",
                        chunk_index=ci + 1,
                        total_chunks=n_chunks,
                    )
                    jobs.append(job)
                    seq += 1
            else:
                job = ExtractionJob(
                    plan_id=f"{run_id}_job_{seq:03d}",
                    sequence_number=seq,
                    discipline=discipline,
                    extraction_scope=cat,
                    categories=[cat],
                    estimated_element_count=count,
                    strategy="isolated_category",
                    reason=(
                        f"{cat} has {count} elements, exceeds "
                        f"isolate_category_threshold ({isolate_category_threshold})"
                    ),
                    expected_prompt=f"Run InventoryModel for {cat}",
                    risk_level="medium",
                )
                jobs.append(job)
                seq += 1

        # Handle small categories — group by discipline
        if small_cats:
            group: list[tuple[str, int]] = []
            group_total = 0

            for cat, count in small_cats:
                if group_total + count > max_group_elements and group:
                    # Flush current group
                    cat_names = [c for c, _ in group]
                    job = ExtractionJob(
                        plan_id=f"{run_id}_job_{seq:03d}",
                        sequence_number=seq,
                        discipline=discipline,
                        extraction_scope=f"{discipline} group",
                        categories=cat_names,
                        estimated_element_count=group_total,
                        strategy="discipline_group",
                        reason=f"{len(cat_names)} small categories grouped under {discipline}",
                        expected_prompt=f"Run InventoryModel for {discipline}",
                        risk_level="low",
                    )
                    jobs.append(job)
                    seq += 1
                    group = []
                    group_total = 0

                group.append((cat, count))
                group_total += count

            if group:
                cat_names = [c for c, _ in group]
                job = ExtractionJob(
                    plan_id=f"{run_id}_job_{seq:03d}",
                    sequence_number=seq,
                    discipline=discipline,
                    extraction_scope=f"{discipline} group",
                    categories=cat_names,
                    estimated_element_count=group_total,
                    strategy="discipline_group",
                    reason=f"{len(cat_names)} small categories grouped under {discipline}",
                    expected_prompt=f"Run InventoryModel for {discipline}",
                    risk_level="low",
                )
                jobs.append(job)
                seq += 1

    plan.jobs = jobs

    # Warnings
    total = sum(category_counts.values())
    if total > 20000:
        plan.warnings.append(
            f"Model has {total} total elements. Full extraction may be slow or crash Revit."
        )
    large_risk = [j for j in jobs if j.risk_level == "high"]
    if large_risk:
        plan.warnings.append(
            f"{len(large_risk)} extraction job(s) have high risk level. "
            "Consider running these last or reducing chunk sizes."
        )

    # Recommend schema discovery before value extraction
    plan.warnings.append(
        "Recommended: run 'Run InventoryModel schema' for whole-model object/class inventory "
        "(ElementId, Category, ClassName, no parameters — validated safe on Revit 2027)."
    )
    plan.warnings.append(
        "For parameter definitions, use category-constrained parameter schema "
        "(e.g. 'Run InventoryModel for Walls parameter schema'). "
        "Whole-model parameter schema is blocked — it crashed Revit 2027."
    )
    if total > 5000:
        plan.warnings.append(
            "For value sampling, use category-constrained sample values "
            "(e.g. 'Run InventoryModel sample values for Walls max 25'). "
            "Whole-model sample values is blocked — it crashed Revit 2027."
        )
    plan.warnings.append(
        "Full value extraction ('Run full InventoryModel') is blocked. "
        "Never recommend full value extraction for live Revit sessions."
    )

    return plan


# ── Plan output writers ──────────────────────────────────────────────


def write_plan_json(plan: ExtractionPlan, path: Path) -> Path:
    """Write the extraction plan as JSON."""
    path.parent.mkdir(parents=True, exist_ok=True)
    data = {
        "run_id": plan.run_id,
        "source_model": plan.source_model,
        "total_instance_count": plan.total_instance_count,
        "total_type_count": plan.total_type_count,
        "total_category_count": plan.total_category_count,
        "thresholds": plan.thresholds,
        "discipline_totals": plan.discipline_totals,
        "categories_by_count": plan.categories_by_count,
        "warnings": plan.warnings,
        "jobs": [asdict(j) for j in plan.jobs],
    }
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    return path


def write_plan_md(plan: ExtractionPlan, path: Path) -> Path:
    """Write the extraction plan as human-readable Markdown."""
    path.parent.mkdir(parents=True, exist_ok=True)

    total_elements = sum(c for _, c in plan.categories_by_count)
    recommended_first = plan.jobs[0] if plan.jobs else None

    lines = [
        f"# Extraction Plan: {plan.run_id}",
        "",
        f"**Source model:** {plan.source_model}",
        f"**Total instances:** {plan.total_instance_count:,}",
        f"**Total types:** {plan.total_type_count:,}",
        f"**Total elements (from categories):** {total_elements:,}",
        f"**Total categories:** {plan.total_category_count}",
        "",
        "## Thresholds",
        "",
        "| Parameter | Value |",
        "|-----------|-------|",
        f"| max_group_elements | {plan.thresholds.get('max_group_elements', 'N/A')} |",
        f"| isolate_category_threshold | {plan.thresholds.get('isolate_category_threshold', 'N/A')} |",
        f"| max_category_chunk_elements | {plan.thresholds.get('max_category_chunk_elements', 'N/A')} |",
        "",
    ]

    # Warnings
    if plan.warnings:
        lines.extend(["## Warnings", ""])
        for w in plan.warnings:
            lines.append(f"- {w}")
        lines.append("")

    # Discipline totals
    lines.extend([
        "## Discipline Totals",
        "",
        "| Discipline | Estimated Elements |",
        "|------------|-------------------|",
    ])
    for disc in DISCIPLINES:
        count = plan.discipline_totals.get(disc, 0)
        lines.append(f"| {disc} | {count:,} |")
    lines.append("")

    # Categories by count (top 30)
    lines.extend([
        "## Categories by Count (top 30)",
        "",
        "| # | Category | Count | Discipline |",
        "|---|----------|-------|------------|",
    ])
    for i, (cat, count) in enumerate(plan.categories_by_count[:30], 1):
        disc = _classify_category_to_discipline(cat)
        lines.append(f"| {i} | {cat} | {count:,} | {disc} |")
    lines.append("")

    # Proposed extraction jobs
    lines.extend([
        f"## Proposed Extraction Jobs ({len(plan.jobs)} total)",
        "",
        "| # | Discipline | Scope | Strategy | Est. Elements | Risk | Prompt |",
        "|---|------------|-------|----------|---------------|------|--------|",
    ])
    for j in plan.jobs:
        lines.append(
            f"| {j.sequence_number} | {j.discipline} | {j.extraction_scope} "
            f"| {j.strategy} | {j.estimated_element_count:,} "
            f"| {j.risk_level} | `{j.expected_prompt}` |"
        )
    lines.append("")

    # Largest risky categories
    risky = [j for j in plan.jobs if j.risk_level in ("medium", "high")]
    if risky:
        lines.extend(["## Largest / Risky Categories", ""])
        for j in risky:
            lines.append(
                f"- **{j.extraction_scope}** ({j.discipline}): "
                f"~{j.estimated_element_count:,} elements, "
                f"strategy={j.strategy}, risk={j.risk_level}"
            )
        lines.append("")

    # Recommended first job
    if recommended_first:
        lines.extend([
            "## Recommended First Extraction",
            "",
            f"Start with **Job #{recommended_first.sequence_number}**: "
            f"{recommended_first.extraction_scope} ({recommended_first.discipline})",
            f"- Estimated elements: {recommended_first.estimated_element_count:,}",
            f"- Strategy: {recommended_first.strategy}",
            f"- Risk: {recommended_first.risk_level}",
            f"- Prompt: `{recommended_first.expected_prompt}`",
            "",
        ])

    # Schema discovery recommendation
    lines.extend([
        "## Recommended: Schema Discovery First",
        "",
        "### Step 1: Object Schema (element/class inventory)",
        "```",
        "Run InventoryModel schema",
        "```",
        "Collects: ElementId, Category, ClassName, Name, LevelName, IsType.",
        "**No parameters.** Safe for whole-model scans. Validated on Revit 2027.",
        "",
        "### Step 2: Parameter Schema (per category — whole-model is blocked)",
        "```",
        "Run InventoryModel for Walls parameter schema",
        "Run InventoryModel for Ceilings parameter schema",
        "Run InventoryModel for Plumbing Fixtures parameter schema",
        "Run InventoryModel parameter schema on Level 1",
        "```",
        "Collects: ParameterName, StorageType, BuiltInParameterId, IsReadOnly, "
        "Instance/Type, ObservedCount, ObservedOnCategories. **No values.**",
        "",
        "**Whole-model parameter schema is blocked** (crashed Revit 2027). "
        "Use category or level-constrained parameter schema.",
        "",
        "### Step 3: Constrained Value Sampling",
        "```",
        "Run InventoryModel sample values for Walls",
        "Run InventoryModel sample values for Walls max 25",
        "Run InventoryModel sample values on Level 1 max 25",
        "```",
        "Hard caps: MaxElements=25, SampleLimit=5 per parameter.",
        "",
        "**Whole-model value sampling is blocked** (crashed Revit 2027). "
        "Always constrain with category, level, or max.",
        "",
        "**Full value extraction ('Run full InventoryModel') is blocked.** "
        "Use category scans for small categories only.",
        "",
    ])

    # Fallback instructions
    lines.extend([
        "## Fallback if a Job Crashes",
        "",
        "1. Note which job number failed.",
        "2. Completed outputs from prior jobs are preserved.",
        "3. Reduce `max_category_chunk_elements` to create smaller chunks.",
        "4. Re-run the planner with adjusted thresholds.",
        "5. Or extract the failing category one sub-category at a time.",
        "",
    ])

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return path


def write_plan_xlsx(plan: ExtractionPlan, path: Path) -> Path:
    """Write the extraction plan as a formatted Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # ── Jobs sheet ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Extraction Jobs"

    headers = [
        "Seq", "Plan ID", "Discipline", "Scope", "Categories",
        "Est. Elements", "Strategy", "Reason", "Prompt", "Risk",
        "Chunk", "Total Chunks",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, j in enumerate(plan.jobs, 2):
        ws.cell(row=row_idx, column=1, value=j.sequence_number)
        ws.cell(row=row_idx, column=2, value=j.plan_id)
        ws.cell(row=row_idx, column=3, value=j.discipline)
        ws.cell(row=row_idx, column=4, value=j.extraction_scope)
        ws.cell(row=row_idx, column=5, value=", ".join(j.categories))
        ws.cell(row=row_idx, column=6, value=j.estimated_element_count)
        ws.cell(row=row_idx, column=7, value=j.strategy)
        ws.cell(row=row_idx, column=8, value=j.reason)
        ws.cell(row=row_idx, column=9, value=j.expected_prompt)
        ws.cell(row=row_idx, column=10, value=j.risk_level)
        ws.cell(row=row_idx, column=11, value=j.chunk_index or "")
        ws.cell(row=row_idx, column=12, value=j.total_chunks or "")

    ws.freeze_panes = "A2"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # ── Categories sheet ─────────────────────────────────────────────
    ws_cat = wb.create_sheet("Categories")
    cat_headers = ["Category", "Count", "Discipline"]
    for col, h in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, (cat, count) in enumerate(plan.categories_by_count, 2):
        ws_cat.cell(row=row_idx, column=1, value=cat)
        ws_cat.cell(row=row_idx, column=2, value=count)
        ws_cat.cell(row=row_idx, column=3, value=_classify_category_to_discipline(cat))

    ws_cat.freeze_panes = "A2"
    ws_cat.column_dimensions["A"].width = 35
    ws_cat.column_dimensions["B"].width = 15
    ws_cat.column_dimensions["C"].width = 20

    # ── Summary sheet ────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    summary_data = [
        ("Metric", "Value"),
        ("Run ID", plan.run_id),
        ("Source Model", plan.source_model),
        ("Total Instances", plan.total_instance_count),
        ("Total Types", plan.total_type_count),
        ("Total Categories", plan.total_category_count),
        ("Extraction Jobs", len(plan.jobs)),
        ("max_group_elements", plan.thresholds.get("max_group_elements", "")),
        ("isolate_category_threshold", plan.thresholds.get("isolate_category_threshold", "")),
        ("max_category_chunk_elements", plan.thresholds.get("max_category_chunk_elements", "")),
    ]
    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_l = ws_sum.cell(row=row_idx, column=1, value=label)
        cell_v = ws_sum.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            cell_l.font = header_font
            cell_l.fill = header_fill
            cell_v.font = header_font
            cell_v.fill = header_fill

    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 30
    ws_sum.freeze_panes = "A2"

    wb.save(str(path))
    return path


PRIORITY_CATEGORIES = [
    "Walls", "Doors", "Windows", "Floors", "Rooms", "Views", "Sheets",
    "Levels", "Grids", "Ducts", "Pipes", "Mechanical Equipment",
    "Plumbing Fixtures", "Lighting Fixtures", "Electrical Fixtures",
    "Ceilings", "Columns", "Stairs", "Railings", "Furniture",
]

BLOCKED_COMMANDS = [
    "Run InventoryModel parameter schema",
    "Run InventoryModel sample values",
    "Run full InventoryModel",
]

# Categories that are not real element categories or cannot be scanned
SKIP_CATEGORIES = {
    "(No Category)",
    "No Category",
    "<Unnamed>",
}


def build_parameter_schema_plan(
    category_counts: dict[str, int],
    run_id: str = "",
    source_model: str = "",
) -> ExtractionPlan:
    """Build a parameter schema discovery plan from summary-mode category counts.

    Produces category-by-category parameter schema commands. Priority
    categories come first, then remaining sorted smallest-to-largest.
    Never recommends whole-model parameter schema.
    Skips non-executable categories (No Category, etc.).
    """
    plan = ExtractionPlan(
        run_id=run_id,
        source_model=source_model,
        total_category_count=len(category_counts),
        thresholds={"mode": "parameter-schema"},
    )

    plan.categories_by_count = sorted(category_counts.items(), key=lambda x: -x[1])

    # Separate priority vs non-priority, filtering out non-executable
    priority_lower = {p.lower(): p for p in PRIORITY_CATEGORIES}
    priority_cats: list[tuple[str, int]] = []
    other_cats: list[tuple[str, int]] = []
    skipped_cats: list[str] = []

    for cat, count in category_counts.items():
        if count <= 0:
            continue
        if cat in SKIP_CATEGORIES:
            skipped_cats.append(cat)
            continue
        if cat.lower() in priority_lower:
            priority_cats.append((cat, count))
        else:
            other_cats.append((cat, count))

    # Priority categories in defined order, others smallest-to-largest
    priority_order = {p.lower(): i for i, p in enumerate(PRIORITY_CATEGORIES)}
    priority_cats.sort(key=lambda x: priority_order.get(x[0].lower(), 999))
    other_cats.sort(key=lambda x: x[1])

    ordered = priority_cats + other_cats

    seq = 1
    jobs: list[ExtractionJob] = []
    total_elements = 0

    for cat, count in ordered:
        risk = "low" if count < 1000 else ("medium" if count < 5000 else "high")
        is_priority = cat.lower() in priority_lower
        job = ExtractionJob(
            plan_id=run_id,
            sequence_number=seq,
            discipline=_classify_category_to_discipline(cat),
            extraction_scope=f"{cat} parameter schema",
            categories=[cat],
            estimated_element_count=count,
            strategy="category_parameter_schema",
            reason=(
                f"PRIORITY: Parameter definitions for {cat} ({count:,} elements)"
                if is_priority
                else f"Parameter definitions for {cat} ({count:,} elements)"
            ),
            expected_prompt=f"Run InventoryModel for {cat} parameter schema",
            risk_level=risk,
        )
        jobs.append(job)
        total_elements += count
        seq += 1

    plan.jobs = jobs
    plan.total_instance_count = total_elements

    plan.warnings.append(
        "Whole-model parameter schema is BLOCKED (crashed Revit 2027). "
        "This plan uses category-by-category extraction."
    )
    plan.warnings.append(
        "BLOCKED unsafe commands (do NOT use):\n"
        + "\n".join(f"    - {cmd}" for cmd in BLOCKED_COMMANDS)
    )
    plan.warnings.append(
        f"Priority categories listed first ({len(priority_cats)} of "
        f"{len(ordered)}), then remaining by size."
    )
    if skipped_cats:
        plan.warnings.append(
            f"Skipped {len(skipped_cats)} non-executable categories: "
            + ", ".join(skipped_cats)
        )
    plan.warnings.append(
        "After completing parameter schema discovery, use:\n"
        "    1. axiom inventory-import-batch --dir <exports-dir> "
        "--scan-mode category_parameter_schema\n"
        "    2. axiom parameter-registry-build "
        "--from-inventory artifacts/model_inventory_runs"
    )

    return plan


def write_parameter_schema_plan_md(plan: ExtractionPlan, path: Path) -> Path:
    """Write parameter schema plan as copyable Markdown with Revit prompts."""
    path.parent.mkdir(parents=True, exist_ok=True)

    priority_lower = {p.lower() for p in PRIORITY_CATEGORIES}
    priority_jobs = [j for j in plan.jobs if j.categories[0].lower() in priority_lower]
    other_jobs = [j for j in plan.jobs if j.categories[0].lower() not in priority_lower]

    lines = [
        f"# Parameter Schema Discovery Plan: {plan.run_id}",
        "",
        f"**Source model:** {plan.source_model}",
        f"**Categories:** {plan.total_category_count}",
        f"**Total jobs:** {len(plan.jobs)} "
        f"({len(priority_jobs)} priority, {len(other_jobs)} remaining)",
        "**Mode:** category-by-category parameter schema",
        "",
        "## Warnings",
        "",
    ]
    for w in plan.warnings:
        lines.append(f"- {w}")
    lines.append("")

    lines.extend([
        "## BLOCKED Commands (do NOT use)",
        "",
    ])
    for cmd in BLOCKED_COMMANDS:
        lines.append(f"- ~~`{cmd}`~~ — crashes Revit or unsafe")
    lines.append("")

    # Priority prompts section
    if priority_jobs:
        lines.extend([
            "## Priority Prompts (copy-paste ready)",
            "",
            "```",
        ])
        for j in priority_jobs:
            lines.append(j.expected_prompt)
        lines.extend([
            "```",
            "",
        ])

    # Remaining prompts
    if other_jobs:
        lines.extend([
            "## Remaining Prompts (smallest first)",
            "",
            "```",
        ])
        for j in other_jobs:
            lines.append(j.expected_prompt)
        lines.extend([
            "```",
            "",
        ])

    # All prompts combined for convenience
    lines.extend([
        "## All Prompts (execution order)",
        "",
        "```",
    ])
    for j in plan.jobs:
        lines.append(j.expected_prompt)
    lines.extend([
        "```",
        "",
        "## Extraction Jobs",
        "",
        "| # | Category | Est. Elements | Risk | Priority | Prompt |",
        "|---|----------|---------------|------|----------|--------|",
    ])
    for j in plan.jobs:
        is_pri = "✓" if j.categories[0].lower() in priority_lower else ""
        lines.append(
            f"| {j.sequence_number} | {j.extraction_scope} "
            f"| {j.estimated_element_count:,} | {j.risk_level} "
            f"| {is_pri} | `{j.expected_prompt}` |"
        )
    lines.append("")

    lines.extend([
        "## Preferred Execution Path",
        "",
        "> **Manual prompt execution is no longer intended for full category coverage.**",
        "> Use the plan execution queue in Revit instead:",
        "",
        "```",
        "Run InventoryModel parameter schema plan max 10",
        "Run InventoryModel parameter schema plan priority only",
        "Run InventoryModel parameter schema plan resume",
        "Run InventoryModel parameter schema plan",
        "```",
        "",
        "## After Completion",
        "",
        "```bash",
        "# Import from manifest (preferred):",
        'axiom inventory-import-batch --manifest "<manifest_path>"',
        "",
        "# Or import from directory:",
        'axiom inventory-import-batch --dir "<exports-dir>" '
        '--scan-mode category_parameter_schema',
        "",
        "# Build registry (with object registry for coverage analysis):",
        "axiom parameter-registry-build --from-inventory artifacts/model_inventory_runs "
        "--object-registry artifacts/object_registry_candidates/<run_id>",
        "",
        "# Review coverage:",
        "axiom inventory-plan-status",
        "```",
        "",
    ])

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return path


def generate_plan_outputs(
    plan: ExtractionPlan,
    output_dir: Path,
    mode: str = "extraction",
) -> dict[str, Path]:
    """Write all plan output files and return paths."""
    plan_dir = output_dir / plan.run_id
    plan_dir.mkdir(parents=True, exist_ok=True)

    paths: dict[str, Path] = {}
    if mode == "parameter-schema":
        paths["json"] = write_plan_json(plan, plan_dir / "parameter_schema_plan.json")
        paths["markdown"] = write_parameter_schema_plan_md(
            plan, plan_dir / "parameter_schema_plan.md",
        )
    else:
        paths["json"] = write_plan_json(plan, plan_dir / "inventory_extraction_plan.json")
        paths["markdown"] = write_plan_md(plan, plan_dir / "inventory_extraction_plan.md")
    paths["xlsx"] = write_plan_xlsx(plan, plan_dir / "inventory_extraction_plan.xlsx")

    return paths
