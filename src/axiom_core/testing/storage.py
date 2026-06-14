"""Layered storage for grid test harness runs.

Writes results to multiple formats:
  1. JSONL — raw append-only event log
  2. SQLite — queryable local execution history (prompt_executions table)
  3. Parquet — durable structured datasets for regression analysis
  4. CSV — human-readable flat export (review output)
  5. XLSX — formatted Excel workbook with Results + Summary sheets (review output)
  6. Markdown — review summary with pass/fail counts and failure details
"""

import csv
import json
from pathlib import Path
from typing import Optional

import pyarrow as pa
import pyarrow.parquet as pq
from sqlalchemy.orm import sessionmaker

from axiom_core.testing.models import GridTestResult

# ── Parquet schema ───────────────────────────────────────────────────
_PARQUET_SCHEMA = pa.schema([
    ("test_id", pa.string()),
    ("prompt", pa.string()),
    ("mode", pa.string()),
    ("git_commit", pa.string()),
    ("git_branch", pa.string()),
    ("timestamp", pa.string()),
    ("resolved_capability", pa.string()),
    ("resolved_parameters", pa.string()),  # JSON string
    ("assumptions", pa.string()),  # JSON string
    ("pipe_available", pa.bool_()),
    ("status", pa.string()),
    ("created_count", pa.int32()),
    ("created_ids", pa.string()),  # JSON string
    ("warnings", pa.string()),  # JSON string
    ("errors", pa.string()),  # JSON string
    ("duration_ms", pa.int32()),
    ("expected_success", pa.bool_()),
    ("expected_created_count", pa.int32()),
    ("expected_capability", pa.string()),
    ("expected_parameters", pa.string()),  # JSON string
    ("passed", pa.bool_()),
    ("failure_category", pa.string()),
    ("failure_detail", pa.string()),
    ("notes", pa.string()),
])


def _result_to_dict(r: GridTestResult) -> dict:
    """Convert a GridTestResult to a flat dict for serialization."""
    return {
        "test_id": r.test_id,
        "prompt": r.prompt,
        "mode": r.mode,
        "git_commit": r.git_commit,
        "git_branch": r.git_branch,
        "timestamp": r.timestamp,
        "resolved_capability": r.resolved_capability or "",
        "resolved_parameters": json.dumps(r.resolved_parameters, default=str),
        "assumptions": json.dumps(r.assumptions),
        "pipe_available": r.pipe_available,
        "status": r.status,
        "created_count": r.created_count,
        "created_ids": json.dumps(r.created_ids),
        "warnings": json.dumps(r.warnings),
        "errors": json.dumps(r.errors),
        "duration_ms": r.duration_ms,
        "expected_success": r.expected_success,
        "expected_created_count": r.expected_created_count,
        "expected_capability": r.expected_capability or "",
        "expected_parameters": json.dumps(r.expected_parameters, default=str),
        "passed": r.passed,
        "failure_category": r.failure_category,
        "failure_detail": r.failure_detail,
        "notes": r.notes,
    }


def write_jsonl(results: list[GridTestResult], path: Path) -> Path:
    """Append results to a JSONL file."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "a", encoding="utf-8") as f:
        for r in results:
            f.write(json.dumps(_result_to_dict(r), default=str) + "\n")
    return path


def write_parquet(results: list[GridTestResult], path: Path) -> Path:
    """Write results to a Parquet file."""
    path.parent.mkdir(parents=True, exist_ok=True)

    rows = [_result_to_dict(r) for r in results]
    arrays = {}
    for field in _PARQUET_SCHEMA:
        col_name = field.name
        values = [row[col_name] for row in rows]
        arrays[col_name] = values

    table = pa.table(arrays, schema=_PARQUET_SCHEMA)
    pq.write_table(table, str(path))
    return path


def read_parquet(path: Path) -> dict[str, list]:
    """Read a Parquet file and return column-oriented dict."""
    if not path.exists():
        return {}
    table = pq.read_table(str(path))
    return table.to_pydict()


def write_to_sqlite(
    results: list[GridTestResult],
    session_factory: Optional[sessionmaker] = None,
) -> None:
    """Persist test results to the SQLite prompt_executions table."""
    if session_factory is None:
        return

    try:
        from axiom_core.database import get_session
        from axiom_core.models import PromptExecutionRow

        with get_session(session_factory) as session:
            for r in results:
                row = PromptExecutionRow(
                    prompt=r.prompt,
                    mode=f"test_{r.mode}",
                    capability=r.resolved_capability,
                    status=r.status,
                    created_count=r.created_count,
                    duration_ms=r.duration_ms,
                )
                row.set_parameters(r.resolved_parameters)
                row.set_assumptions(r.assumptions)
                row.set_created_ids(r.created_ids)
                row.set_errors(r.errors)
                row.set_warnings(r.warnings)
                session.add(row)
    except Exception:
        pass


def persist_results(
    results: list[GridTestResult],
    output_dir: Path,
    run_id: str,
    session_factory: Optional[sessionmaker] = None,
) -> dict[str, Path]:
    """Write results to all three storage layers.

    Returns dict of {format: path} for the files written.
    """
    run_dir = output_dir / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    paths: dict[str, Path] = {}

    paths["jsonl"] = write_jsonl(results, run_dir / "results.jsonl")
    paths["parquet"] = write_parquet(results, run_dir / "results.parquet")

    write_to_sqlite(results, session_factory)

    return paths


# ── Review export column order ───────────────────────────────────────
# Key columns placed first for human review, remaining appended after.
_REVIEW_COLUMNS_PRIORITY = [
    "test_id",
    "prompt",
    "resolved_capability",
    "status",
    "passed",
    "created_count",
    "expected_created_count",
    "failure_category",
    "failure_detail",
    "warnings",
    "errors",
    "resolved_parameters",
    "notes",
    "mode",
    "git_branch",
    "git_commit",
    "timestamp",
    "duration_ms",
]


def _review_rows(results: list[GridTestResult]) -> list[dict]:
    """Build flat dicts in review column order."""
    raw = [_result_to_dict(r) for r in results]
    if not raw:
        return []
    # Build ordered rows: priority columns first, then any remaining
    all_keys = list(raw[0].keys())
    remaining = [k for k in all_keys if k not in _REVIEW_COLUMNS_PRIORITY]
    ordered = [k for k in _REVIEW_COLUMNS_PRIORITY if k in all_keys] + remaining
    return [{k: row[k] for k in ordered} for row in raw]


def write_csv(results: list[GridTestResult], path: Path) -> Path:
    """Write results to a CSV file with review column ordering."""
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = _review_rows(results)
    if not rows:
        path.write_text("", encoding="utf-8")
        return path
    fieldnames = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return path


def write_xlsx(results: list[GridTestResult], path: Path, run_id: str = "") -> Path:
    """Write results to a formatted Excel workbook with Results + Summary sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    rows = _review_rows(results)
    wb = Workbook()

    # ── Results sheet ────────────────────────────────────────────────
    ws_results = wb.active
    ws_results.title = "Results"

    if rows:
        headers = list(rows[0].keys())
    else:
        headers = list(_REVIEW_COLUMNS_PRIORITY)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws_results.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row.get(header, "")
            if isinstance(val, bool):
                val = str(val)
            cell = ws_results.cell(row=row_idx, column=col_idx, value=val)
            # Highlight pass/fail in the "passed" column
            if header == "passed":
                cell.fill = pass_fill if val == "True" else fail_fill

    # Freeze header row
    ws_results.freeze_panes = "A2"

    # Autosize columns (approximate: min of max content width, capped at 50)
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row_idx in range(2, len(rows) + 2):
            cell_val = str(ws_results.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, min(len(cell_val), 60))
        ws_results.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    # ── Summary sheet ────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")

    total = len(results)
    passed = sum(1 for r in results if r.passed)
    failed = sum(1 for r in results if not r.passed and r.failure_category != "skipped")
    clarification = sum(1 for r in results if r.status == "CLARIFICATION_NEEDED")
    skipped = sum(1 for r in results if r.failure_category == "skipped")

    summary_data = [
        ("Metric", "Count"),
        ("Run ID", run_id),
        ("Total scenarios", total),
        ("Passed", passed),
        ("Failed", failed),
        ("Clarification needed", clarification),
        ("Skipped", skipped),
    ]
    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_label = ws_summary.cell(row=row_idx, column=1, value=label)
        cell_value = ws_summary.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            cell_label.font = header_font
            cell_label.fill = header_fill
            cell_value.font = header_font
            cell_value.fill = header_fill

    # Failed tests list
    failed_results = [
        r for r in results if not r.passed and r.failure_category != "skipped"
    ]
    if failed_results:
        start_row = len(summary_data) + 2
        ws_summary.cell(row=start_row, column=1, value="Failed Scenarios").font = Font(bold=True)
        ws_summary.cell(row=start_row + 1, column=1, value="Test ID").font = header_font
        ws_summary.cell(row=start_row + 1, column=2, value="Prompt").font = header_font
        ws_summary.cell(row=start_row + 1, column=3, value="Failure Reason").font = header_font
        for i, r in enumerate(failed_results):
            ws_summary.cell(row=start_row + 2 + i, column=1, value=r.test_id)
            ws_summary.cell(row=start_row + 2 + i, column=2, value=r.prompt[:80])
            ws_summary.cell(row=start_row + 2 + i, column=3, value=r.failure_detail[:100])

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 40
    ws_summary.column_dimensions["C"].width = 50
    ws_summary.freeze_panes = "A2"

    wb.save(str(path))
    return path


def write_review_summary_md(
    results: list[GridTestResult], path: Path, run_id: str = "",
    output_paths: Optional[dict[str, Path]] = None,
) -> Path:
    """Write a concise markdown review summary."""
    path.parent.mkdir(parents=True, exist_ok=True)

    total = len(results)
    passed = sum(1 for r in results if r.passed)
    failed = sum(1 for r in results if not r.passed and r.failure_category != "skipped")
    clarification = sum(1 for r in results if r.status == "CLARIFICATION_NEEDED")
    skipped = sum(1 for r in results if r.failure_category == "skipped")

    lines = [
        f"# Scenario Results Summary: {run_id}",
        "",
        "## Counts",
        "",
        "| Metric | Count |",
        "|--------|-------|",
        f"| Total scenarios | {total} |",
        f"| Passed | {passed} |",
        f"| Failed | {failed} |",
        f"| Clarification needed | {clarification} |",
        f"| Skipped | {skipped} |",
        "",
    ]

    # Failed tests
    failed_results = [
        r for r in results if not r.passed and r.failure_category != "skipped"
    ]
    if failed_results:
        lines.extend(["## Failures", ""])
        for r in failed_results:
            lines.append(
                f"- **{r.test_id}**: {r.failure_detail or r.failure_category}"
            )
        lines.append("")

    # Clarification-needed tests
    clar_results = [r for r in results if r.status == "CLARIFICATION_NEEDED"]
    if clar_results:
        lines.extend(["## Clarification Needed", ""])
        for r in clar_results:
            lines.append(f"- **{r.test_id}**: `{r.prompt[:80]}`")
        lines.append("")

    # Output files
    if output_paths:
        lines.extend(["## Output Files", ""])
        for fmt, p in output_paths.items():
            lines.append(f"- **{fmt}:** `{p}`")
        lines.append("")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return path


def export_review_package(
    results: list[GridTestResult],
    run_dir: Path,
    run_id: str = "",
) -> dict[str, Path]:
    """Export a complete human-review package (CSV + XLSX + Markdown summary).

    Returns dict of {format: path} for the review files written.
    """
    run_dir.mkdir(parents=True, exist_ok=True)
    paths: dict[str, Path] = {}

    paths["csv"] = write_csv(results, run_dir / "scenario_results.csv")
    paths["xlsx"] = write_xlsx(results, run_dir / "scenario_results.xlsx", run_id=run_id)

    # Markdown summary written last so it can include output paths
    all_paths = dict(paths)
    all_paths["parquet"] = run_dir / "results.parquet"
    all_paths["jsonl"] = run_dir / "results.jsonl"

    paths["summary_md"] = write_review_summary_md(
        results, run_dir / "scenario_results_summary.md",
        run_id=run_id, output_paths=all_paths,
    )

    return paths
