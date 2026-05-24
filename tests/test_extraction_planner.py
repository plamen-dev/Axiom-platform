"""Tests for the adaptive extraction planner."""

import json
import tempfile
from pathlib import Path

from axiom_core.inventory.extraction_planner import (
    build_extraction_plan,
    generate_plan_outputs,
)


def _small_model_counts():
    """Category counts for a small model — everything fits in discipline groups."""
    return {
        "Doors": 500,
        "Windows": 300,
        "Furniture": 700,
        "Rooms": 200,
        "Structural Columns": 150,
        "Structural Framing": 400,
        "Ducts": 800,
        "Duct Fittings": 200,
        "Air Terminals": 150,
        "Conduits": 300,
        "Lighting Fixtures": 600,
        "Electrical Fixtures": 200,
        "Pipes": 400,
        "Plumbing Fixtures": 100,
        "Sprinklers": 50,
        "Generic Models": 80,
    }


def _large_model_counts():
    """Category counts for a large model — requires isolation and chunking."""
    return {
        "Walls": 2500,
        "Doors": 500,
        "Windows": 300,
        "Furniture": 700,
        "Rooms": 200,
        "Ceilings": 400,
        "Floors": 800,
        "Structural Columns": 1500,
        "Structural Framing": 2000,
        "Structural Foundations": 300,
        "Ducts": 7500,
        "Duct Fittings": 1200,
        "Air Terminals": 900,
        "Mechanical Equipment": 150,
        "Conduits": 2800,
        "Lighting Fixtures": 1800,
        "Electrical Fixtures": 500,
        "Electrical Equipment": 100,
        "Fire Alarm Devices": 400,
        "Data Devices": 300,
        "Pipes": 20000,
        "Pipe Fittings": 2500,
        "Plumbing Fixtures": 800,
        "Sprinklers": 1200,
        "Generic Models": 250,
    }


class TestSmallCategoryGrouping:
    """Test 1: Small categories are grouped by discipline."""

    def test_small_categories_grouped_by_discipline(self):
        plan = build_extraction_plan(_small_model_counts(), run_id="test_small")

        # All jobs should be discipline_group strategy
        strategies = {j.strategy for j in plan.jobs}
        assert strategies == {"discipline_group"}

        # Every discipline with categories should have a job
        disciplines_with_jobs = {j.discipline for j in plan.jobs}
        assert "Architectural" in disciplines_with_jobs
        assert "Structural" in disciplines_with_jobs
        assert "Mechanical" in disciplines_with_jobs
        assert "Electrical" in disciplines_with_jobs
        assert "Plumbing" in disciplines_with_jobs
        assert "Other" in disciplines_with_jobs


class TestDisciplineSplitByLargestCategories:
    """Test 2: A discipline over the threshold is split by largest categories."""

    def test_large_discipline_splits_large_categories(self):
        plan = build_extraction_plan(_large_model_counts(), run_id="test_split")

        # Mechanical discipline has Ducts=7500 which exceeds isolate threshold
        mech_jobs = [j for j in plan.jobs if j.discipline == "Mechanical"]
        assert len(mech_jobs) > 1  # Should have isolated Ducts + grouped small ones

        # Ducts should be isolated or chunked, not in a discipline_group
        duct_jobs = [j for j in mech_jobs if "Ducts" in j.categories]
        assert len(duct_jobs) >= 1
        assert duct_jobs[0].strategy in ("isolated_category", "category_chunk")


class TestSingleLargeCategoryIsolated:
    """Test 3: A single large category is isolated."""

    def test_large_category_isolated(self):
        counts = {"Pipes": 4000, "Plumbing Fixtures": 200, "Sprinklers": 100}
        plan = build_extraction_plan(counts, run_id="test_isolate")

        pipe_jobs = [j for j in plan.jobs if "Pipes" in j.categories]
        assert len(pipe_jobs) == 1
        assert pipe_jobs[0].strategy == "isolated_category"
        assert pipe_jobs[0].estimated_element_count == 4000


class TestVeryLargeCategoryChunked:
    """Test 4: A very large category is chunked."""

    def test_very_large_category_chunked(self):
        counts = {"Walls": 12000}
        plan = build_extraction_plan(
            counts, run_id="test_chunk",
            max_category_chunk_elements=5000,
        )

        wall_jobs = [j for j in plan.jobs if "Walls" in j.categories]
        assert len(wall_jobs) == 3  # 12000 / 5000 = 2.4, rounds to 3
        assert all(j.strategy == "category_chunk" for j in wall_jobs)
        assert wall_jobs[0].chunk_index == 1
        assert wall_jobs[0].total_chunks == 3


class TestPipes20kChunking:
    """Test 5: Pipes with 20,000 elements become multiple chunks."""

    def test_pipes_20k_chunked(self):
        counts = {"Pipes": 20000, "Plumbing Fixtures": 500}
        plan = build_extraction_plan(
            counts, run_id="test_pipes",
            max_category_chunk_elements=5000,
        )

        pipe_jobs = [j for j in plan.jobs if "Pipes" in j.categories]
        assert len(pipe_jobs) == 4  # 20000 / 5000 = 4
        assert all(j.strategy == "category_chunk" for j in pipe_jobs)
        assert all(j.discipline == "Plumbing" for j in pipe_jobs)

        total_estimated = sum(j.estimated_element_count for j in pipe_jobs)
        assert total_estimated == 20000


class TestDucts7500Chunking:
    """Test 6: Ducts with 7,500 elements become multiple chunks."""

    def test_ducts_7500_chunked(self):
        counts = {"Ducts": 7500, "Air Terminals": 200}
        plan = build_extraction_plan(
            counts, run_id="test_ducts",
            max_category_chunk_elements=5000,
        )

        duct_jobs = [j for j in plan.jobs if "Ducts" in j.categories]
        assert len(duct_jobs) == 2  # 7500 / 5000 = 1.5, rounds to 2
        assert all(j.strategy == "category_chunk" for j in duct_jobs)
        assert all(j.discipline == "Mechanical" for j in duct_jobs)


class TestGenericModelsGoToOther:
    """Test 7: Generic Models go to Other and are not forced into a discipline."""

    def test_generic_models_classified_as_other(self):
        counts = {"Generic Models": 500, "Walls": 1000}
        plan = build_extraction_plan(counts, run_id="test_generic")

        generic_jobs = [j for j in plan.jobs if "Generic Models" in j.categories]
        assert len(generic_jobs) >= 1
        assert all(j.discipline == "Other" for j in generic_jobs)


class TestPlanOutputFilesCreated:
    """Test 8: Plan output JSON/Markdown/XLSX are created."""

    def test_plan_outputs_exist(self):
        plan = build_extraction_plan(
            _small_model_counts(),
            run_id="test_outputs",
            source_model="Test Model",
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            paths = generate_plan_outputs(plan, Path(tmpdir))

            assert paths["json"].exists()
            assert paths["markdown"].exists()
            assert paths["xlsx"].exists()

            assert paths["json"].name == "inventory_extraction_plan.json"
            assert paths["markdown"].name == "inventory_extraction_plan.md"
            assert paths["xlsx"].name == "inventory_extraction_plan.xlsx"

            # Verify JSON content
            data = json.loads(paths["json"].read_text())
            assert data["run_id"] == "test_outputs"
            assert data["source_model"] == "Test Model"
            assert len(data["jobs"]) > 0

            # Verify Markdown has key sections
            md = paths["markdown"].read_text()
            assert "Extraction Plan" in md
            assert "Thresholds" in md
            assert "Discipline Totals" in md
            assert "Proposed Extraction Jobs" in md

            # Verify XLSX has sheets
            from openpyxl import load_workbook
            wb = load_workbook(str(paths["xlsx"]))
            assert "Extraction Jobs" in wb.sheetnames
            assert "Categories" in wb.sheetnames
            assert "Summary" in wb.sheetnames


class TestSummaryModeJsonAccepted:
    """Test 9: Summary-mode JSON with empty elements but category counts is accepted."""

    def test_summary_mode_with_category_counts(self):
        # This simulates what the Revit summary export looks like
        plan = build_extraction_plan(
            category_counts={"Walls": 5000, "Ducts": 3000, "Pipes": 2000},
            run_id="test_summary",
            source_model="Summary Model",
            total_instance_count=10000,
            total_type_count=500,
        )

        assert plan.total_instance_count == 10000
        assert plan.total_type_count == 500
        assert len(plan.jobs) > 0
        assert plan.source_model == "Summary Model"


class TestPlannerDoesNotRequireElements:
    """Test 10: Planner does not require full-detail elements."""

    def test_planner_works_without_elements(self):
        # Planner takes category_counts dict, not elements list
        counts = {"Walls": 1000, "Doors": 500, "Pipes": 8000}
        plan = build_extraction_plan(counts, run_id="test_no_elements")

        # Should produce valid plan
        assert len(plan.jobs) > 0
        assert plan.total_category_count == 3

        # Pipes should be chunked
        pipe_jobs = [j for j in plan.jobs if "Pipes" in j.categories]
        assert len(pipe_jobs) >= 1
        assert pipe_jobs[0].strategy in ("isolated_category", "category_chunk")

        with tempfile.TemporaryDirectory() as tmpdir:
            paths = generate_plan_outputs(plan, Path(tmpdir))
            assert all(p.exists() for p in paths.values())
