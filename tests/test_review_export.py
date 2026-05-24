"""Tests for human-readable review export (CSV, XLSX, Markdown)."""

import csv
import tempfile
from pathlib import Path

from axiom_core.testing.models import GridTestResult
from axiom_core.testing.storage import (
    export_review_package,
    write_csv,
    write_review_summary_md,
    write_xlsx,
)


def _make_results() -> list[GridTestResult]:
    """Build a small set of mixed test results for export testing."""
    return [
        GridTestResult(
            test_id="grid_uniform_10x10",
            prompt="Create 10 vertical grids spaced 10 feet apart",
            mode="simulate",
            git_commit="abc123",
            git_branch="main",
            timestamp="2026-05-19T12:00:00Z",
            resolved_capability="CreateGrids",
            resolved_parameters={"VerticalCount": 10, "SpacingFeet": 10.0},
            status="SUCCESS",
            created_count=10,
            expected_success=True,
            expected_created_count=10,
            passed=True,
            duration_ms=5,
        ),
        GridTestResult(
            test_id="grid_arithmetic_spacing",
            prompt="Create 10 grids spaced 5, 10, 15 and so on",
            mode="simulate",
            git_commit="abc123",
            git_branch="main",
            timestamp="2026-05-19T12:00:01Z",
            resolved_capability="CreateGrids",
            resolved_parameters={},
            status="CLARIFICATION_NEEDED",
            created_count=0,
            expected_success=True,
            expected_created_count=0,
            expected_capability="CreateGrids",
            passed=True,
            duration_ms=2,
            notes="Arithmetic spacing triggers clarification",
        ),
        GridTestResult(
            test_id="grid_wrong_count",
            prompt="Create 5 vertical grids spaced 8 feet apart",
            mode="simulate",
            git_commit="abc123",
            git_branch="main",
            timestamp="2026-05-19T12:00:02Z",
            resolved_capability="CreateGrids",
            resolved_parameters={"VerticalCount": 5, "SpacingFeet": 8.0},
            status="SUCCESS",
            created_count=4,
            expected_success=True,
            expected_created_count=5,
            passed=False,
            failure_category="count_mismatch",
            failure_detail="Expected 5 elements, got 4",
            duration_ms=4,
        ),
        GridTestResult(
            test_id="grid_skipped_real",
            prompt="Create grids in real mode",
            mode="real",
            git_commit="abc123",
            git_branch="main",
            timestamp="2026-05-19T12:00:03Z",
            status="SKIPPED",
            created_count=0,
            expected_success=True,
            expected_created_count=0,
            passed=False,
            failure_category="skipped",
            failure_detail="Skipped: real mode not available",
            duration_ms=0,
        ),
    ]


class TestWriteCSV:
    def test_csv_created(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "results.csv"
            write_csv(_make_results(), path)
            assert path.exists()
            assert path.stat().st_size > 0

    def test_csv_has_header_and_rows(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "results.csv"
            results = _make_results()
            write_csv(results, path)

            with open(path, newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                rows = list(reader)

            assert len(rows) == len(results)
            # First column should be test_id (priority order)
            assert list(rows[0].keys())[0] == "test_id"

    def test_csv_column_order(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "results.csv"
            write_csv(_make_results(), path)

            with open(path, newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                header = next(reader)

            # Key columns should come first
            assert header[0] == "test_id"
            assert header[1] == "prompt"
            assert "passed" in header[:8]
            assert "status" in header[:8]


class TestWriteXLSX:
    def test_xlsx_created(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "results.xlsx"
            write_xlsx(_make_results(), path, run_id="test_run")
            assert path.exists()
            assert path.stat().st_size > 0

    def test_xlsx_has_results_and_summary_sheets(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "results.xlsx"
            write_xlsx(_make_results(), path, run_id="test_run")

            wb = load_workbook(str(path))
            sheet_names = wb.sheetnames

            assert "Results" in sheet_names
            assert "Summary" in sheet_names

    def test_xlsx_results_sheet_has_frozen_header(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "results.xlsx"
            write_xlsx(_make_results(), path, run_id="test_run")

            wb = load_workbook(str(path))
            ws = wb["Results"]
            assert ws.freeze_panes == "A2"

    def test_xlsx_results_row_count(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "results.xlsx"
            results = _make_results()
            write_xlsx(results, path, run_id="test_run")

            wb = load_workbook(str(path))
            ws = wb["Results"]
            # Header row + data rows
            assert ws.max_row == 1 + len(results)

    def test_xlsx_summary_counts_match(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "results.xlsx"
            results = _make_results()
            write_xlsx(results, path, run_id="test_run")

            wb = load_workbook(str(path))
            ws = wb["Summary"]

            # Read summary metrics into a dict
            summary = {}
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[0] is not None:
                    summary[row[0]] = row[1]

            total = len(results)
            passed = sum(1 for r in results if r.passed)
            failed = sum(
                1 for r in results
                if not r.passed and r.failure_category != "skipped"
            )
            clarification = sum(
                1 for r in results if r.status == "CLARIFICATION_NEEDED"
            )
            skipped = sum(
                1 for r in results if r.failure_category == "skipped"
            )

            assert summary["Total scenarios"] == total
            assert summary["Passed"] == passed
            assert summary["Failed"] == failed
            assert summary["Clarification needed"] == clarification
            assert summary["Skipped"] == skipped


class TestWriteReviewSummaryMd:
    def test_markdown_created(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "summary.md"
            write_review_summary_md(_make_results(), path, run_id="test_run")
            assert path.exists()
            assert path.stat().st_size > 0

    def test_markdown_contains_counts(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "summary.md"
            results = _make_results()
            write_review_summary_md(results, path, run_id="test_run")

            content = path.read_text(encoding="utf-8")
            assert "Total scenarios" in content
            assert "Passed" in content
            assert "Failed" in content
            assert "Clarification needed" in content

    def test_markdown_lists_failures(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "summary.md"
            write_review_summary_md(_make_results(), path, run_id="test_run")

            content = path.read_text(encoding="utf-8")
            assert "grid_wrong_count" in content
            assert "Failures" in content

    def test_markdown_lists_clarification(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "summary.md"
            write_review_summary_md(_make_results(), path, run_id="test_run")

            content = path.read_text(encoding="utf-8")
            assert "grid_arithmetic_spacing" in content
            assert "Clarification Needed" in content

    def test_markdown_includes_output_paths(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "summary.md"
            output_paths = {
                "csv": Path(tmpdir) / "results.csv",
                "xlsx": Path(tmpdir) / "results.xlsx",
            }
            write_review_summary_md(
                _make_results(), path, run_id="test_run",
                output_paths=output_paths,
            )

            content = path.read_text(encoding="utf-8")
            assert "Output Files" in content
            assert "csv" in content
            assert "xlsx" in content


class TestExportReviewPackage:
    def test_all_files_created(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            run_dir = Path(tmpdir) / "run_001"
            paths = export_review_package(
                _make_results(), run_dir, run_id="run_001",
            )

            assert "csv" in paths
            assert "xlsx" in paths
            assert "summary_md" in paths

            assert paths["csv"].exists()
            assert paths["xlsx"].exists()
            assert paths["summary_md"].exists()

            assert paths["csv"].name == "scenario_results.csv"
            assert paths["xlsx"].name == "scenario_results.xlsx"
            assert paths["summary_md"].name == "scenario_results_summary.md"

    def test_empty_results(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            run_dir = Path(tmpdir) / "run_empty"
            paths = export_review_package([], run_dir, run_id="run_empty")

            assert paths["csv"].exists()
            assert paths["xlsx"].exists()
            assert paths["summary_md"].exists()
